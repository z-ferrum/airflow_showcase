import os, csv
from openpyxl import load_workbook  # to be tested that all required data in a real dataset is read properly

from airflow import DAG
from airflow.providers.snowflake.operators.snowflake import SnowflakeOperator
from airflow.operators.python_operator import PythonOperator
from airflow.utils.dates import days_ago



# INPUTS

excel_data_folder = 'micro_data'  # micro_data for 10-line .xlsx files, data for full .xlsx files
dag_script_directory = os.path.dirname(os.path.abspath(__file__))  # current file directory
base_directory = os.path.join(dag_script_directory, '..', excel_data_folder)  # base directory relatively to the file

table_names = [
    'device',
    'store',
    'transaction'
]

default_args = {
    'owner': 'admin',
    'email_on_failure': False,
    'email_on_retry': False,
    'depends_on_past': False,   # since we do one-time analysis that is manually triggered. 
                                # for an incrementally updated pipeline True would be better.
}

common_snowflake_args = {
    'snowflake_conn_id': 'snowflake_conn',  # connection is configured in Airflow web UI
    'database': 'SHOWCASE',
    'schema': 'RAW'
}

copy_tasks = []


# FUNCTIONS

def convert_excel_to_csv(excel_path,csv_path):

    excel_file = load_workbook(excel_path)               # loads the workbook and selects the first worksheet 
    data_sheet = excel_file.worksheets[0]               # selects first sheet in the excel file

    with open(csv_path, 'w', encoding='utf-8', newline='') as csv_file:   # Opens the CSV file
        writer = csv.writer(csv_file)
        
        for row in data_sheet.iter_rows():              # Iterates through the rows in the worksheet and write to the CSV file
            row_data = [cell.value for cell in row]
            writer.writerow(row_data)


def delete_file(file_path):  # isn't critical, should not break the pipeline
    try:
        os.remove(file_path)
        print(f"File {file_path} has been deleted.")
    except FileNotFoundError:
        print(f"File {file_path} not found.")
    except PermissionError:
        print(f"Permission denied for deleting {file_path}.")
    except Exception as e:
        print(f"An error occurred: {e}")



# DAGS

with DAG(
    'excel_to_snowflake',
    default_args=default_args,
    description='Converts Excel files into CSVs, loads to Snowflake and ',
    schedule_interval=None,  # Manual trigger, since we consider 1-time analysis
    start_date=days_ago(1),
    catchup=False  # Since we consider 1-time analysis
) as dag:

    for table_name in table_names:

        csv_file_name = table_name + '.csv'
        excel_file_name = table_name + '.xlsx'
        csv_path = os.path.join(base_directory, csv_file_name)
        excel_path = os.path.join(base_directory, excel_file_name)  

        convert_to_csv = PythonOperator (
            task_id=f'convert_{excel_file_name}_to_csv',
            python_callable=convert_excel_to_csv,
            op_args=[excel_path,csv_path]
        )

        stage_sql = f'PUT file://{csv_path} @BACK_STAGE AUTO_COMPRESS=TRUE'
        
        stage_task = SnowflakeOperator(
            task_id=f'stage_{csv_file_name}_to_snowflake',
            sql=stage_sql,
            **common_snowflake_args
        )

        copy_sql = f"""
            TRUNCATE {table_name};  
            COPY INTO {table_name}
            FROM '@BACK_STAGE/{os.path.basename(csv_path)}'
            FILE_FORMAT = (TYPE = 'CSV' FIELD_OPTIONALLY_ENCLOSED_BY = '"')
            ON_ERROR = 'CONTINUE';
            """

        copy_task = SnowflakeOperator(
            task_id=f'copy_{table_name}_to_warehouse',
            sql=copy_sql,
            **common_snowflake_args
        )

        copy_tasks.append(copy_task)

        delete_csv = PythonOperator (               # With large data and multiple files from a partitioned datalake
            task_id=f'delete_{csv_file_name}.csv',  # we wouldn't want to keep both .xlsx and .csv copies
            python_callable=delete_file,            # opt for keeping the original file
            op_args=[csv_path]
        )


        # Task order
        convert_to_csv >> stage_task >> copy_task >> [delete_csv.as_teardown()]  # csv deleted last to save compute resources if we need to re-run the pipeline
        
        # end of copy_tasks loop


    clean_stage = SnowflakeOperator(
        task_id=f'clean_stage_{csv_file_name}',
        sql='REMOVE @back_Stage/',  # deletes the files from the stage area
        **common_snowflake_args
    )
    
    for copy_task in copy_tasks:
        copy_task >> [clean_stage.as_teardown()]